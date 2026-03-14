"""Validation report structures and export helpers."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from sqlalchemy_excel._compat import sanitize_cell_value

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class CellError:
    """Represents one cell-level validation error.

    Attributes:
        row: Excel row number (1-based).
        column: Logical column name or header text.
        value: Original raw value that failed validation.
        expected_type: Human-readable expected type description.
        message: Human-readable validation message.
        error_code: Machine-readable code for categorizing errors.
    """

    row: int
    column: str
    value: Any  # pyright: ignore[reportExplicitAny]
    expected_type: str
    message: str
    error_code: str


@dataclass
class ValidationReport:
    """Aggregated validation results for a single worksheet/file.

    Attributes:
        errors: Collected cell-level errors.
        total_rows: Total number of data rows validated (excluding header row).
        valid_rows: Number of rows with no validation errors.
        invalid_rows: Number of rows containing at least one validation error.
    """

    errors: list[CellError]
    total_rows: int
    valid_rows: int
    invalid_rows: int

    @property
    def has_errors(self) -> bool:
        """Return ``True`` when at least one validation error exists."""

        return bool(self.errors)

    def summary(self) -> str:
        """Return a compact human-readable validation summary."""

        return (
            f"Validated {self.total_rows} rows: {self.valid_rows} valid, "
            f"{self.invalid_rows} invalid. {len(self.errors)} errors found."
        )

    def to_dict(self) -> dict[str, object]:
        """Return a JSON-serializable dictionary representation."""

        return {
            "errors": [asdict(error) for error in self.errors],
            "total_rows": self.total_rows,
            "valid_rows": self.valid_rows,
            "invalid_rows": self.invalid_rows,
            "has_errors": self.has_errors,
            "summary": self.summary(),
        }

    def errors_by_row(self) -> dict[int, list[CellError]]:
        """Group all collected errors by Excel row number."""

        grouped: dict[int, list[CellError]] = defaultdict(list)
        for error in self.errors:
            grouped[error.row].append(error)
        return dict(grouped)

    def to_excel(self, path: str | Path) -> None:
        """Export validation errors to an Excel workbook.

        Args:
            path: Output path for the generated `.xlsx` report.
        """

        workbook = Workbook()
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            worksheet = cast("Worksheet", workbook.create_sheet("Validation Errors"))
        worksheet.title = "Validation Errors"

        headers = ["row", "column", "value", "expected_type", "message", "error_code"]
        _ = worksheet.append(headers)

        error_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )

        for error in self.errors:
            value_obj = cast("object", error.value)
            if value_obj is None:
                rendered_value: object = None
            elif isinstance(value_obj, str):
                rendered_value = sanitize_cell_value(value_obj)
            else:
                rendered_value = str(value_obj)
            _ = worksheet.append(
                [
                    error.row,
                    error.column,
                    rendered_value,
                    error.expected_type,
                    error.message,
                    error.error_code,
                ]
            )
            row_index = worksheet.max_row
            for cell in worksheet[row_index]:
                cell.fill = error_fill

        workbook.save(Path(path))
