"""Validation orchestration for Excel sources."""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING, BinaryIO, Protocol, cast

from openpyxl import load_workbook

from sqlalchemy_excel.validation.pydantic_backend import PydanticBackend
from sqlalchemy_excel.validation.report import CellError, ValidationReport

if TYPE_CHECKING:
    from collections.abc import Iterable, Mapping
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

    from sqlalchemy_excel.mapping import ExcelMapping
    from sqlalchemy_excel.reader.openpyxl_reader import OpenpyxlReader


@dataclass
class _ReaderResult:
    headers: list[str]
    rows: Iterable[Mapping[str, object]]
    total_rows: int | None


class _ReaderProtocol(Protocol):
    def read(
        self,
        source: str | Path | BinaryIO,
        sheet_name: str | None = None,
        header_row: int = 1,
    ) -> _ReaderResult: ...


_ExternalOpenpyxlReader: type[OpenpyxlReader] | None
try:
    from sqlalchemy_excel.reader.openpyxl_reader import (
        OpenpyxlReader as _ExternalOpenpyxlReader,
    )
except ImportError:
    _ExternalOpenpyxlReader = None


class ExcelValidator:
    """Orchestrate row-level validation for Excel files."""

    def __init__(self, mappings: list[ExcelMapping], *, backend: str = "pydantic") -> None:
        """Initialize validator with mapping metadata and backend selection.

        Args:
            mappings: Mapping definitions, one per worksheet schema.
            backend: Validation backend identifier. Currently supports ``pydantic``.
        """

        if not mappings:
            raise ValueError("At least one mapping is required")
        if backend != "pydantic":
            raise ValueError("Only 'pydantic' backend is currently supported")

        self._mappings: list[ExcelMapping] = mappings
        self._backends: dict[str, PydanticBackend] = {
            mapping.sheet_name: PydanticBackend(mapping) for mapping in mappings
        }
        self._reader: _ReaderProtocol = _build_reader()

    def validate(
        self,
        source: str | Path | BinaryIO,
        *,
        sheet_name: str | None = None,
        max_errors: int | None = None,
        stop_on_first_error: bool = False,
    ) -> ValidationReport:
        """Validate Excel rows and return a structured report.

        Args:
            source: Input Excel file path or binary stream.
            sheet_name: Optional explicit worksheet name.
            max_errors: Optional cap on the number of collected errors.
            stop_on_first_error: Stop processing after first invalid row.

        Returns:
            A ``ValidationReport`` containing row counts and cell-level errors.
        """

        mapping = _select_mapping(self._mappings, sheet_name)
        backend = self._backends[mapping.sheet_name]
        reader_result = self._reader.read(source, sheet_name=sheet_name, header_row=1)
        header_map = _build_header_map(mapping, reader_result.headers)

        errors: list[CellError] = []
        invalid_rows = 0
        processed_rows = 0

        for offset, raw_row in enumerate(reader_result.rows, start=2):
            processed_rows += 1
            row_number = offset
            row_data = _remap_row(raw_row, header_map, mapping)
            row_errors = backend.validate_row(row_data, row_number)
            if row_errors:
                invalid_rows += 1
                errors.extend(row_errors)

                if max_errors is not None and len(errors) >= max_errors:
                    errors = errors[:max_errors]
                    break
                if stop_on_first_error:
                    break

        total_rows = reader_result.total_rows if reader_result.total_rows is not None else processed_rows
        valid_rows = max(total_rows - invalid_rows, 0)
        return ValidationReport(
            errors=errors,
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
        )


def _select_mapping(mappings: list[ExcelMapping], sheet_name: str | None) -> ExcelMapping:
    if sheet_name is None:
        return mappings[0]

    for mapping in mappings:
        if mapping.sheet_name == sheet_name:
            return mapping

    return mappings[0]


def _build_reader() -> _ReaderProtocol:
    if _ExternalOpenpyxlReader is not None:
        return cast("_ReaderProtocol", cast("object", _ExternalOpenpyxlReader()))
    return _LocalOpenpyxlReader()


class _LocalOpenpyxlReader:
    def read(
        self,
        source: str | Path | BinaryIO,
        sheet_name: str | None = None,
        header_row: int = 1,
    ) -> _ReaderResult:
        workbook = load_workbook(filename=source, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        worksheet = cast("Worksheet | None", worksheet)
        if worksheet is None:
            raise ValueError("Workbook does not contain a readable worksheet")

        first_row = next(
            worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True),
            None,
        )
        headers = [str(cell).strip() if cell is not None else "" for cell in (first_row or ())]

        rows: list[Mapping[str, object]] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            row: dict[str, object] = {
                headers[index]: cast("object", value)
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            rows.append(row)

        return _ReaderResult(headers=headers, rows=rows, total_rows=len(rows))


def _normalize_header(value: str) -> str:
    return "_".join(value.strip().lower().split())


def _build_header_map(mapping: ExcelMapping, headers: list[str]) -> dict[str, str]:
    normalized_to_column: dict[str, str] = {}
    for column in mapping.columns:
        normalized_to_column[_normalize_header(column.name)] = column.name
        normalized_to_column[_normalize_header(column.excel_header)] = column.name

    return {
        header: normalized_to_column.get(_normalize_header(header), header)
        for header in headers
        if header
    }


def _remap_row(
    raw_row: Mapping[str, object],
    header_map: dict[str, str],
    mapping: ExcelMapping,
) -> dict[str, object]:
    remapped: dict[str, object] = {column.name: None for column in mapping.columns}
    for header, value in raw_row.items():
        target_column = header_map.get(header)
        if target_column in remapped:
            remapped[target_column] = value
    return remapped
