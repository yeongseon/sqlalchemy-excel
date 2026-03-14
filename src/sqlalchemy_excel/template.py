"""Excel template generation from Excel mappings."""

from __future__ import annotations

import enum
import os
import tempfile
from contextlib import suppress
from datetime import date, datetime
from io import BytesIO
from typing import TYPE_CHECKING, cast

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from sqlalchemy_excel._compat import sanitize_cell_value
from sqlalchemy_excel.excelio.session import ExcelWorkbookSession
from sqlalchemy_excel.exceptions import TemplateError

if TYPE_CHECKING:
    from collections.abc import Callable, Sequence
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

    from sqlalchemy_excel.mapping import ColumnMapping, ExcelMapping

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HEADER_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_MAX_EXCEL_ROWS = 1_048_576
_MAX_EXCEL_DATA_VALIDATION_FORMULA_LENGTH = 255
_WIDTH_PADDING = 2
CellValue = str | int | float | bool | date | datetime | None


class ExcelTemplate:
    """Generate formatted Excel templates from one or more mappings.

    Args:
        mappings: Collection of model-to-sheet mappings.
        include_sample_data: If true, include one representative data row.
    """

    def __init__(
        self,
        mappings: list[ExcelMapping],
        *,
        include_sample_data: bool = False,
    ) -> None:
        self._mappings: list[ExcelMapping] = mappings
        self._include_sample_data: bool = include_sample_data

    def save(self, path: str | Path) -> None:
        """Save the generated template workbook to disk.

        Args:
            path: Destination .xlsx path.

        Raises:
            TemplateError: If workbook generation or saving fails.
        """
        try:
            with ExcelWorkbookSession.open(path, create=True) as session:
                self._populate_workbook(session.workbook)
                session.commit()
        except Exception as exc:
            raise TemplateError(f"Failed to save template to '{path}': {exc}") from exc

    def to_bytes(self) -> bytes:
        """Render the template workbook to bytes.

        Returns:
            XLSX bytes.

        Raises:
            TemplateError: If workbook generation fails.
        """
        buffer = self.to_bytesio()
        return buffer.getvalue()

    def to_bytesio(self) -> BytesIO:
        """Render the template workbook to an in-memory stream.

        Returns:
            BytesIO containing XLSX data.

        Raises:
            TemplateError: If workbook generation fails.
        """
        temp_path: str | None = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                temp_path = tmp.name

            self.save(temp_path)

            with open(temp_path, "rb") as file_obj:
                stream = BytesIO(file_obj.read())
        except Exception as exc:
            raise TemplateError(f"Failed to render template workbook: {exc}") from exc
        finally:
            if temp_path is not None:
                with suppress(OSError):
                    os.unlink(temp_path)

        _ = stream.seek(0)
        return stream

    def _build_workbook(self) -> Workbook:
        workbook = Workbook()
        self._populate_workbook(workbook)
        return workbook

    def _populate_workbook(self, workbook: Workbook) -> None:
        if not self._mappings:
            raise TemplateError("At least one mapping is required to build a template")

        for sheet_name in list(workbook.sheetnames):
            del workbook[sheet_name]

        try:
            for mapping in self._mappings:
                self._add_sheet(workbook, mapping)
        except Exception as exc:
            raise TemplateError(f"Failed to build template workbook: {exc}") from exc

    def _add_sheet(self, workbook: Workbook, mapping: ExcelMapping) -> None:
        worksheet = cast("Worksheet", workbook.create_sheet(title=mapping.sheet_name))

        for index, column in enumerate(mapping.columns, start=1):
            cell = worksheet.cell(row=1, column=index, value=column.excel_header)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER
            cell.fill = REQUIRED_FILL if self._is_required(column) else HEADER_FILL
            cell.comment = Comment(self._build_comment_text(column), "sqlalchemy-excel")

            self._set_column_width(worksheet, index, column)

            if column.enum_values:
                self._add_enum_validation(worksheet, index, column)

        last_column_letter = get_column_letter(len(mapping.columns))
        worksheet.auto_filter.ref = f"A1:{last_column_letter}1"
        worksheet.freeze_panes = "A2"

        if self._include_sample_data:
            self._add_sample_row(worksheet, mapping.columns)

    def _build_comment_text(self, column: ColumnMapping) -> str:
        parts: list[str] = [f"Type: {self._column_type_hint(column)}"]

        if column.max_length is not None:
            parts.append(f"Max length: {column.max_length}")

        parts.append("Required" if self._is_required(column) else "Optional")

        if column.primary_key:
            parts.append("Primary key")

        if column.foreign_key:
            parts.append(f"Foreign key: {column.foreign_key}")

        if column.description:
            parts.append(f"Description: {column.description}")

        return ", ".join(parts)

    def _set_column_width(
        self,
        worksheet: Worksheet,
        index: int,
        column: ColumnMapping,
    ) -> None:
        type_hint_length = len(self._column_type_hint(column))
        header_length = len(column.excel_header)
        width = max(header_length, type_hint_length) + _WIDTH_PADDING
        worksheet.column_dimensions[get_column_letter(index)].width = float(width)

    def _add_enum_validation(
        self,
        worksheet: Worksheet,
        index: int,
        column: ColumnMapping,
    ) -> None:
        values = column.enum_values
        if not values:
            return

        escaped_values = [value.replace('"', '""') for value in values]
        formula = f'"{",".join(escaped_values)}"'
        if len(formula) > _MAX_EXCEL_DATA_VALIDATION_FORMULA_LENGTH:
            header_cell = worksheet.cell(row=1, column=index)
            note = (
                "Dropdown omitted: enum list exceeds Excel data-validation "
                "255-character limit."
            )
            existing_comment = (
                header_cell.comment.text if header_cell.comment is not None else ""
            )
            comment_text = f"{existing_comment}, {note}" if existing_comment else note
            header_cell.comment = Comment(comment_text, "sqlalchemy-excel")
            return

        validation = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=column.nullable,
            showDropDown=False,
        )
        worksheet.add_data_validation(validation)

        letter = get_column_letter(index)
        validation_range = f"{letter}2:{letter}{_MAX_EXCEL_ROWS}"
        add_validation_range = cast("Callable[[str], None]", validation.add)
        add_validation_range(validation_range)

    def _add_sample_row(
        self,
        worksheet: Worksheet,
        columns: Sequence[ColumnMapping],
    ) -> None:
        sample_row = 2
        for index, column in enumerate(columns, start=1):
            value = self._sample_value(column)
            _ = worksheet.cell(row=sample_row, column=index, value=value)

    def _sample_value(self, column: ColumnMapping) -> CellValue:
        if column.enum_values:
            return sanitize_cell_value(column.enum_values[0])

        py_type = column.python_type

        if py_type is int:
            return 1
        if py_type is float:
            return 1.0
        if py_type is str:
            return sanitize_cell_value("sample")
        if py_type is bool:
            return True
        if py_type is date:
            return date.today()
        if py_type is datetime:
            return datetime.now()
        if issubclass(py_type, enum.Enum):
            first_member = next(iter(py_type), None)
            if isinstance(first_member, enum.Enum):
                return first_member.name

        return None

    def _column_type_hint(self, column: ColumnMapping) -> str:
        if column.enum_values:
            return f"enum[{', '.join(column.enum_values)}]"

        py_type = column.python_type
        return py_type.__name__

    def _is_required(self, column: ColumnMapping) -> bool:
        return not column.nullable and not column.has_default
