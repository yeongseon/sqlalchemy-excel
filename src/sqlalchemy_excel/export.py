"""Query result → Excel export."""

from __future__ import annotations

import os
import tempfile
from contextlib import suppress
from datetime import date, datetime
from typing import TYPE_CHECKING, Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy_excel._compat import sanitize_cell_value
from sqlalchemy_excel.excelio.session import ExcelWorkbookSession
from sqlalchemy_excel.exceptions import ExportError

if TYPE_CHECKING:
    from collections.abc import Sequence
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

    from sqlalchemy_excel.mapping import ExcelMapping


class ExcelExporter:
    """Export SQLAlchemy query results to formatted Excel files.

    Args:
        mappings: List of ExcelMapping instances defining the export structure.

    Example:
        >>> exporter = ExcelExporter([mapping])
        >>> exporter.export(users, "users_export.xlsx")
    """

    HEADER_FILL = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # Date/time formats for Excel
    DATE_FORMAT = "YYYY-MM-DD"
    DATETIME_FORMAT = "YYYY-MM-DD HH:MM:SS"

    def __init__(self, mappings: list[ExcelMapping]) -> None:
        if not mappings:
            raise ExportError("At least one ExcelMapping is required")
        self._mappings = mappings

    def export(
        self,
        rows: Sequence[Any],
        path: str | Path | None = None,
        *,
        sheet_name: str | None = None,
    ) -> bytes | None:
        """Export rows to an Excel file.

        Args:
            rows: Sequence of ORM model instances or dicts.
            path: File path to save. If None, returns bytes.
            sheet_name: Override sheet name from mapping.

        Returns:
            Excel file as bytes if path is None, otherwise None.

        Raises:
            ExportError: If export fails.
        """
        if path is not None:
            try:
                with ExcelWorkbookSession.open(path, create=True) as session:
                    self._populate_workbook(session.workbook, rows, sheet_name)
                    session.commit()
            except Exception as e:
                raise ExportError(f"Failed to save Excel file: {e}") from e
            return None

        temp_path: str | None = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                temp_path = tmp.name

            _ = self.export(rows, temp_path, sheet_name=sheet_name)

            with open(temp_path, "rb") as file_obj:
                return file_obj.read()
        except Exception as e:
            raise ExportError(f"Failed to write Excel to buffer: {e}") from e
        finally:
            if temp_path is not None:
                with suppress(OSError):
                    os.unlink(temp_path)

    def _populate_workbook(
        self,
        wb: Workbook,
        rows: Sequence[Any],
        sheet_name_override: str | None,
    ) -> None:
        """Populate an openpyxl Workbook with formatted data.

        Args:
            wb: Workbook to populate in place.
            rows: Data rows to export.
            sheet_name_override: Optional sheet name override.
        """

        for sheet_name in list(wb.sheetnames):
            del wb[sheet_name]

        for mapping in self._mappings:
            ws = wb.create_sheet()

            ws.title = sheet_name_override or mapping.sheet_name

            # Write headers
            for col_idx, col_mapping in enumerate(mapping.columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_mapping.excel_header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = self.HEADER_ALIGNMENT

            # Write data rows
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, col_mapping in enumerate(mapping.columns, start=1):
                    value = self._extract_value(row_data, col_mapping.name)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Apply number format for dates
                    if isinstance(value, datetime):
                        cell.number_format = self.DATETIME_FORMAT
                    elif isinstance(value, date):
                        cell.number_format = self.DATE_FORMAT

            # Auto-fit column widths
            for col_idx, col_mapping in enumerate(mapping.columns, start=1):
                col_letter = get_column_letter(col_idx)
                max_width = len(col_mapping.excel_header)

                for row_idx in range(2, len(rows) + 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_width = max(max_width, len(str(cell_value)))

                ws.column_dimensions[col_letter].width = min(max_width + 4, 50)

            # Auto-filter
            if mapping.columns:
                last_col = get_column_letter(len(mapping.columns))
                ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

            # Freeze header row
            ws.freeze_panes = "A2"

    @staticmethod
    def _extract_value(row: Any, column_name: str) -> Any:
        """Extract a value from a row object.

        Supports ORM model instances (getattr) and dicts.

        Args:
            row: The row object (ORM instance or dict).
            column_name: The column/attribute name to extract.

        Returns:
            The extracted value.
        """
        if isinstance(row, dict):
            value = row.get(column_name)
        else:
            value = getattr(row, column_name, None)

        if isinstance(value, str):
            return sanitize_cell_value(value)
        return value
