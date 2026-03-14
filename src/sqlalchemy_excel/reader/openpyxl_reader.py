from __future__ import annotations

import os
from pathlib import Path
from typing import TYPE_CHECKING, BinaryIO, TypeAlias

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from sqlalchemy_excel.exceptions import FileFormatError, ReaderError, SheetNotFoundError
from sqlalchemy_excel.reader.base import ReaderResult, normalize_header

if TYPE_CHECKING:
    from collections.abc import Iterable

    from openpyxl.workbook import Workbook

    from sqlalchemy_excel._types import FileSource, RowDict

WorksheetLike: TypeAlias = Worksheet | ReadOnlyWorksheet


class OpenpyxlReader:
    """Default reader backend implemented with openpyxl."""

    def __init__(
        self,
        *,
        read_only: bool = False,
        max_file_size: int = 50 * 1024 * 1024,
    ) -> None:
        """Initialize an openpyxl-backed Excel reader.

        Args:
            read_only: Whether to open workbook in streaming mode.
            max_file_size: Maximum allowed file size in bytes.
        """

        self.read_only: bool = read_only
        self.max_file_size: int = max_file_size

    def read(
        self,
        source: FileSource,
        sheet_name: str | None = None,
        header_row: int | None = None,
    ) -> ReaderResult:
        """Read rows from an Excel worksheet.

        Args:
            source: File path or binary stream containing an Excel file.
            sheet_name: Optional sheet name. Defaults to workbook active sheet.
            header_row: Optional 1-based header row index. Auto-detected when omitted.

        Returns:
            ReaderResult containing normalized headers and iterable row mappings.

        Raises:
            ReaderError: If file size checks or worksheet parsing fails.
            FileFormatError: If the source is not a valid xlsx file.
            SheetNotFoundError: If the requested sheet does not exist.
        """

        self._validate_file_size(source)

        try:
            workbook = load_workbook(source, read_only=self.read_only)
        except InvalidFileException as exc:
            raise FileFormatError("Input is not a valid .xlsx file") from exc
        except OSError as exc:
            raise ReaderError(f"Unable to open Excel source: {exc}") from exc

        try:
            worksheet = self._select_worksheet(workbook, sheet_name)
            resolved_header_row = self._resolve_header_row(worksheet, header_row)
            headers = self._extract_headers(worksheet, resolved_header_row)

            if self.read_only:
                rows: Iterable[RowDict] = self._iter_rows_streaming(
                    worksheet,
                    headers,
                    resolved_header_row,
                    workbook,
                )
                total_rows = None
            else:
                parsed_rows = tuple(
                    self._iter_row_values(worksheet, headers, resolved_header_row)
                )
                rows = parsed_rows
                total_rows = len(parsed_rows)
                workbook.close()

            return ReaderResult(headers=headers, rows=rows, total_rows=total_rows)
        except Exception:
            workbook.close()
            raise

    def _validate_file_size(self, source: FileSource) -> None:
        if isinstance(source, (str, Path, os.PathLike)):
            self._validate_path_size(source)
            return

        self._validate_stream_size(source)

    def _validate_path_size(self, source: str | os.PathLike[str]) -> None:
        try:
            size = os.path.getsize(source)
        except OSError as exc:
            raise ReaderError(f"Unable to access Excel source: {exc}") from exc

        self._raise_if_exceeds_size(size)

    def _validate_stream_size(self, source: BinaryIO) -> None:
        try:
            original_pos = source.tell()
            _ = source.seek(0, os.SEEK_END)
            size = source.tell()
            _ = source.seek(original_pos)
        except (AttributeError, OSError) as exc:
            raise ReaderError("Unable to determine binary source file size") from exc

        self._raise_if_exceeds_size(size)

    def _raise_if_exceeds_size(self, size: int) -> None:
        if size > self.max_file_size:
            raise ReaderError(
                f"File size ({size:,} bytes) exceeds maximum ({self.max_file_size:,} bytes)"
            )

    def _select_worksheet(self, workbook: Workbook, sheet_name: str | None) -> WorksheetLike:
        if sheet_name is None:
            selected = workbook.active
        else:
            try:
                selected = workbook[sheet_name]
            except KeyError as exc:
                raise SheetNotFoundError(sheet_name, list(workbook.sheetnames)) from exc

        if selected is None:
            raise ReaderError("Selected sheet is not a worksheet")

        return selected

    def _resolve_header_row(self, worksheet: WorksheetLike, header_row: int | None) -> int:
        if header_row is not None:
            if header_row < 1:
                raise ReaderError("header_row must be a positive 1-based index")
            return header_row

        for row_index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if any(not self._is_empty_cell(value) for value in values):
                return row_index

        raise ReaderError("Unable to detect header row: worksheet is empty")

    def _extract_headers(self, worksheet: WorksheetLike, header_row: int) -> list[str]:
        header_values = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ),
            None,
        )
        if header_values is None:
            raise ReaderError(f"Unable to read header row {header_row}")

        last_header_index = 0
        for index, value in enumerate(header_values, start=1):
            if not self._is_empty_cell(value):
                last_header_index = index

        if last_header_index == 0:
            raise ReaderError(f"Header row {header_row} does not contain any values")

        headers: list[str] = []
        for column_index, value in enumerate(header_values[:last_header_index], start=1):
            text = "" if value is None else str(value)
            normalized = normalize_header(text)
            if not normalized:
                raise ReaderError(
                    f"Header value is empty after normalization at column {column_index}"
                )
            if normalized in headers:
                raise ReaderError(f"Duplicate normalized header detected: '{normalized}'")
            headers.append(normalized)

        return headers

    def _iter_rows_streaming(
        self,
        worksheet: WorksheetLike,
        headers: list[str],
        header_row: int,
        workbook: Workbook,
    ) -> Iterable[RowDict]:
        try:
            yield from self._iter_row_values(worksheet, headers, header_row)
        finally:
            workbook.close()

    def _iter_row_values(
        self,
        worksheet: WorksheetLike,
        headers: list[str],
        header_row: int,
    ) -> Iterable[RowDict]:
        header_count = len(headers)
        for row_values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row_values[:header_count])
            if len(values) < header_count:
                values.extend([None] * (header_count - len(values)))

            if all(self._is_empty_cell(value) for value in values):
                continue

            yield dict(zip(headers, values, strict=True))

    @staticmethod
    def _is_empty_cell(value: object) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False
