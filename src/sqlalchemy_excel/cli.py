"""Click CLI entry point for sqlalchemy-excel."""

from __future__ import annotations

import importlib
import json
import sys
from typing import Any

import click

from sqlalchemy_excel.exceptions import SqlalchemyExcelError


def _resolve_model(dotpath: str) -> Any:
    """Resolve a dotted path like 'myapp.models:User' to a class.

    Args:
        dotpath: Module path with colon-separated class name.

    Returns:
        The resolved class.

    Raises:
        click.BadParameter: If resolution fails.
    """
    if ":" not in dotpath:
        raise click.BadParameter(
            f"Model path must be in 'module.path:ClassName' format, got '{dotpath}'"
        )

    module_path, class_name = dotpath.rsplit(":", 1)

    try:
        module = importlib.import_module(module_path)
    except ImportError as e:
        raise click.BadParameter(f"Cannot import module '{module_path}': {e}") from e

    try:
        return getattr(module, class_name)
    except AttributeError as e:
        raise click.BadParameter(
            f"Module '{module_path}' has no class '{class_name}'"
        ) from e


@click.group()
@click.version_option(package_name="sqlalchemy-excel")
def cli() -> None:
    """sqlalchemy-excel: SQLAlchemy model-driven Excel toolkit.

    Generate templates, validate uploads, import to DB, and export from DB.
    """


@cli.command()
@click.option(
    "--model",
    required=True,
    help="ORM model path (e.g., 'myapp.models:User')",
)
@click.option(
    "--output",
    default="template.xlsx",
    help="Output file path",
    show_default=True,
)
@click.option(
    "--sample-data",
    is_flag=True,
    default=False,
    help="Include sample data row",
)
@click.option(
    "--sheet-name",
    default=None,
    help="Override sheet name (default: table name)",
)
def template(
    model: str,
    output: str,
    sample_data: bool,
    sheet_name: str | None,
) -> None:
    """Generate an Excel template from a SQLAlchemy model."""
    from sqlalchemy_excel.mapping import ExcelMapping
    from sqlalchemy_excel.template import ExcelTemplate

    model_class = _resolve_model(model)

    try:
        mapping = ExcelMapping.from_model(
            model_class,
            sheet_name=sheet_name,
        )
        tpl = ExcelTemplate([mapping], include_sample_data=sample_data)
        tpl.save(output)
        click.echo(f"Template saved to {output}")
    except SqlalchemyExcelError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.option(
    "--model",
    required=True,
    help="ORM model path (e.g., 'myapp.models:User')",
)
@click.option(
    "--input",
    "input_path",
    required=True,
    help="Excel file to validate",
)
@click.option(
    "--format",
    "output_format",
    type=click.Choice(["text", "json", "excel"]),
    default="text",
    show_default=True,
    help="Output format for validation report",
)
@click.option(
    "--output",
    default=None,
    help="Output file path (for excel format)",
)
def validate(
    model: str,
    input_path: str,
    output_format: str,
    output: str | None,
) -> None:
    """Validate an Excel file against a SQLAlchemy model."""
    from sqlalchemy_excel.mapping import ExcelMapping
    from sqlalchemy_excel.validation import ExcelValidator

    model_class = _resolve_model(model)

    try:
        mapping = ExcelMapping.from_model(model_class)
        validator = ExcelValidator([mapping])
        report = validator.validate(input_path)

        if output_format == "json":
            click.echo(json.dumps(report.to_dict(), indent=2, default=str))
        elif output_format == "excel":
            out_path = output or "validation_report.xlsx"
            report.to_excel(out_path)
            click.echo(f"Validation report saved to {out_path}")
        else:
            click.echo(report.summary())

        if report.has_errors:
            sys.exit(1)
    except SqlalchemyExcelError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command(name="import")
@click.option(
    "--model",
    required=True,
    help="ORM model path (e.g., 'myapp.models:User')",
)
@click.option(
    "--input",
    "input_path",
    required=True,
    help="Excel file to import",
)
@click.option(
    "--db",
    required=True,
    help="Database URL (e.g., 'sqlite:///app.db')",
)
@click.option(
    "--mode",
    type=click.Choice(["insert", "upsert"]),
    default="insert",
    show_default=True,
    help="Import mode",
)
@click.option(
    "--dry-run",
    is_flag=True,
    default=False,
    help="Preview import without persisting",
)
@click.option(
    "--batch-size",
    default=1000,
    show_default=True,
    help="Batch size for DB operations",
)
def import_cmd(
    model: str,
    input_path: str,
    db: str,
    mode: str,
    dry_run: bool,
    batch_size: int,
) -> None:
    """Import an Excel file into a database."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import Session

    from sqlalchemy_excel.load import ExcelImporter
    from sqlalchemy_excel.mapping import ExcelMapping

    model_class = _resolve_model(model)

    try:
        mapping = ExcelMapping.from_model(model_class)
        engine = create_engine(db)

        with Session(engine) as session:
            importer = ExcelImporter([mapping], session=session)

            if dry_run:
                result = importer.dry_run(input_path)
                click.echo("DRY RUN (no changes persisted):")
            elif mode == "upsert":
                result = importer.upsert(input_path, batch_size=batch_size)
            else:
                result = importer.insert(input_path, batch_size=batch_size)

            if not dry_run:
                session.commit()

            click.echo(result.summary())

    except SqlalchemyExcelError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.option(
    "--model",
    required=True,
    help="ORM model path (e.g., 'myapp.models:User')",
)
@click.option(
    "--db",
    required=True,
    help="Database URL (e.g., 'sqlite:///app.db')",
)
@click.option(
    "--output",
    default="export.xlsx",
    show_default=True,
    help="Output file path",
)
def export(
    model: str,
    db: str,
    output: str,
) -> None:
    """Export database records to an Excel file."""
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from sqlalchemy_excel.export import ExcelExporter
    from sqlalchemy_excel.mapping import ExcelMapping

    model_class = _resolve_model(model)

    try:
        mapping = ExcelMapping.from_model(model_class)
        engine = create_engine(db)

        with Session(engine) as session:
            stmt = select(model_class)
            rows = list(session.execute(stmt).scalars().all())

        exporter = ExcelExporter([mapping])
        exporter.export(rows, output)
        click.echo(f"Exported {len(rows)} rows to {output}")

    except SqlalchemyExcelError as e:
        click.echo(f"Error: {e}", err=True)
        sys.exit(1)


@cli.command()
@click.option(
    "--input",
    "input_path",
    required=True,
    help="Excel file to inspect",
)
def inspect(input_path: str) -> None:
    """Inspect an Excel file's structure (sheets, headers, row counts)."""
    from openpyxl import load_workbook

    from sqlalchemy_excel.exceptions import ReaderError

    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
    except Exception as e:
        raise click.ClickException(f"Cannot open file: {e}") from e

    try:
        click.echo(f"File: {input_path}")
        click.echo(f"Sheets: {len(wb.sheetnames)}")
        click.echo()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            click.echo(f"  Sheet: {sheet_name}")

            # Get headers from first row
            headers: list[str] = []
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row:
                headers = [str(h) if h is not None else "" for h in first_row]
                click.echo(f"  Headers: {', '.join(h for h in headers if h)}")

            # Count rows (approximate for read_only)
            row_count = 0
            for _ in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
            click.echo(f"  Data rows: {row_count}")

            # Show column types (infer from first few data rows)
            if row_count > 0:
                click.echo("  Columns:")
                sample_rows = list(
                    ws.iter_rows(
                        min_row=2,
                        max_row=min(6, row_count + 1),
                        values_only=True,
                    )
                )
                for col_idx, header in enumerate(headers):
                    if not header:
                        continue
                    sample_values = [
                        row[col_idx]
                        for row in sample_rows
                        if col_idx < len(row) and row[col_idx] is not None
                    ]
                    types = {type(v).__name__ for v in sample_values}
                    type_str = ", ".join(sorted(types)) if types else "unknown"
                    click.echo(f"    {header}: {type_str}")

            click.echo()
    except ReaderError:
        raise
    finally:
        wb.close()
