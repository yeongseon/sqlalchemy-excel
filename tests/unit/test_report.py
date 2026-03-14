from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import load_workbook

from sqlalchemy_excel.validation.report import CellError, ValidationReport

if TYPE_CHECKING:
    from pathlib import Path


def test_cell_error_creation() -> None:
    error = CellError(
        row=3,
        column="age",
        value="abc",
        expected_type="int",
        message="Input should be a valid integer",
        error_code="type_error",
    )

    assert error.row == 3
    assert error.column == "age"
    assert error.value == "abc"
    assert error.expected_type == "int"
    assert error.message == "Input should be a valid integer"
    assert error.error_code == "type_error"


def test_validation_report_has_errors_true() -> None:
    report = ValidationReport(
        errors=[
            CellError(
                row=2,
                column="name",
                value=None,
                expected_type="str",
                message="Field required",
                error_code="null_error",
            )
        ],
        total_rows=1,
        valid_rows=0,
        invalid_rows=1,
    )

    assert report.has_errors is True


def test_validation_report_has_errors_false() -> None:
    report = ValidationReport(errors=[], total_rows=2, valid_rows=2, invalid_rows=0)

    assert report.has_errors is False


def test_validation_report_summary() -> None:
    report = ValidationReport(errors=[], total_rows=5, valid_rows=5, invalid_rows=0)

    assert report.summary() == "Validated 5 rows: 5 valid, 0 invalid. 0 errors found."


def test_validation_report_to_dict() -> None:
    report = ValidationReport(
        errors=[
            CellError(
                row=4,
                column="status",
                value="unknown",
                expected_type="one of ['active', 'inactive']",
                message="Input should be 'active' or 'inactive'",
                error_code="enum_error",
            )
        ],
        total_rows=3,
        valid_rows=2,
        invalid_rows=1,
    )

    data = report.to_dict()

    assert set(data.keys()) == {
        "errors",
        "total_rows",
        "valid_rows",
        "invalid_rows",
        "has_errors",
        "summary",
    }
    assert data["total_rows"] == 3
    assert data["valid_rows"] == 2
    assert data["invalid_rows"] == 1
    assert data["has_errors"] is True
    assert isinstance(data["errors"], list)
    assert data["summary"] == "Validated 3 rows: 2 valid, 1 invalid. 1 errors found."


def test_validation_report_errors_by_row() -> None:
    errors = [
        CellError(
            row=2,
            column="name",
            value=None,
            expected_type="str",
            message="Field required",
            error_code="null_error",
        ),
        CellError(
            row=2,
            column="email",
            value="bad",
            expected_type="str",
            message="Invalid email",
            error_code="constraint_error",
        ),
        CellError(
            row=3,
            column="age",
            value="oops",
            expected_type="int",
            message="Input should be a valid integer",
            error_code="type_error",
        ),
    ]
    report = ValidationReport(errors=errors, total_rows=2, valid_rows=0, invalid_rows=2)

    grouped = report.errors_by_row()

    assert set(grouped.keys()) == {2, 3}
    assert [error.column for error in grouped[2]] == ["name", "email"]
    assert [error.column for error in grouped[3]] == ["age"]


def test_validation_report_to_excel(tmp_path: Path) -> None:
    errors = [
        CellError(
            row=2,
            column="name",
            value=None,
            expected_type="str",
            message="Field required",
            error_code="null_error",
        ),
        CellError(
            row=3,
            column="age",
            value="not_a_number",
            expected_type="int",
            message="Input should be a valid integer",
            error_code="type_error",
        ),
    ]
    report = ValidationReport(errors=errors, total_rows=2, valid_rows=0, invalid_rows=2)
    output_path = tmp_path / "validation_report.xlsx"

    report.to_excel(output_path)

    workbook = load_workbook(output_path)
    worksheet = workbook["Validation Errors"]

    assert worksheet.max_row == 3
    assert worksheet.max_column == 6
    assert [cell.value for cell in worksheet[1]] == [
        "row",
        "column",
        "value",
        "expected_type",
        "message",
        "error_code",
    ]
    assert [cell.value for cell in worksheet[2]] == [
        2,
        "name",
        None,
        "str",
        "Field required",
        "null_error",
    ]
    assert [cell.value for cell in worksheet[3]] == [
        3,
        "age",
        "not_a_number",
        "int",
        "Input should be a valid integer",
        "type_error",
    ]
