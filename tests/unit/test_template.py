from __future__ import annotations

# pyright: reportImplicitRelativeImport=none, reportMissingImports=none, reportUnknownVariableType=none, reportUnknownMemberType=none, reportUnknownParameterType=none, reportUnknownArgumentType=none, reportMissingParameterType=none, reportAny=none
from io import BytesIO

import pytest
from conftest import Employee, SimpleUser
from openpyxl import load_workbook
from sqlalchemy import String

from sqlalchemy_excel.exceptions import TemplateError
from sqlalchemy_excel.mapping import ColumnMapping, ExcelMapping
from sqlalchemy_excel.template import ExcelTemplate


def _workbook_from_bytes(data: bytes):
    return load_workbook(BytesIO(data))


def test_save_creates_file(tmp_path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    output = tmp_path / "test.xlsx"

    template.save(output)

    assert output.exists()


def test_to_bytes_returns_bytes() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])

    data = template.to_bytes()

    assert isinstance(data, bytes)
    assert len(data) > 0


def test_to_bytesio_returns_bytesio() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])

    stream = template.to_bytesio()

    assert isinstance(stream, BytesIO)
    assert len(stream.read()) > 0


def test_header_row_values(tmp_path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    output = tmp_path / "headers.xlsx"

    template.save(output)

    workbook = load_workbook(output)
    worksheet = workbook[mapping.sheet_name]
    headers = [
        worksheet.cell(row=1, column=i).value
        for i in range(1, len(mapping.columns) + 1)
    ]

    assert headers == [column.excel_header for column in mapping.columns]


def test_header_formatting() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    allowed = {"4472C4", "FFF2CC"}
    for index in range(1, len(mapping.columns) + 1):
        cell = worksheet.cell(row=1, column=index)
        rgb = cell.fill.start_color.rgb or ""
        assert rgb[-6:] in allowed


def test_freeze_panes() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    assert worksheet.freeze_panes == "A2"


def test_auto_filter() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    assert worksheet.auto_filter.ref is not None


def test_enum_data_validation() -> None:
    mapping = ExcelMapping.from_model(Employee)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    validations = list(worksheet.data_validations.dataValidation)
    assert validations

    status_values = ["active", "inactive", "on_leave"]
    matching_validation = next(
        validation
        for validation in validations
        if all(value in validation.formula1 for value in status_values)
    )
    assert matching_validation.type == "list"


def test_sample_data_included() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping], include_sample_data=True)
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    row_two_values = [
        worksheet.cell(row=2, column=i).value
        for i in range(1, len(mapping.columns) + 1)
    ]
    assert any(value is not None for value in row_two_values)


def test_sample_data_not_included_by_default() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    assert worksheet.max_row == 1


def test_multiple_mappings_multiple_sheets() -> None:
    user_mapping = ExcelMapping.from_model(SimpleUser)
    employee_mapping = ExcelMapping.from_model(Employee)
    template = ExcelTemplate([user_mapping, employee_mapping])
    workbook = _workbook_from_bytes(template.to_bytes())

    assert user_mapping.sheet_name in workbook.sheetnames
    assert employee_mapping.sheet_name in workbook.sheetnames
    assert len(workbook.sheetnames) == 2


def test_empty_mappings_raises_error() -> None:
    template = ExcelTemplate([])

    with pytest.raises(TemplateError, match="At least one mapping is required"):
        template.to_bytes()


def test_column_comments() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template = ExcelTemplate([mapping])
    workbook = _workbook_from_bytes(template.to_bytes())
    worksheet = workbook[mapping.sheet_name]

    for index in range(1, len(mapping.columns) + 1):
        comment = worksheet.cell(row=1, column=index).comment
        assert comment is not None
        assert "Type:" in comment.text


def test_sample_data_sanitizes_enum_formula_values() -> None:
    mapping = ExcelMapping(
        model_class=SimpleUser,
        sheet_name="dangerous_sample",
        columns=[
            ColumnMapping(
                name="name",
                excel_header="Name",
                python_type=str,
                sqla_type=String(),
                nullable=False,
                primary_key=False,
                has_default=False,
                enum_values=["=CMD()"],
            )
        ],
        key_columns=[],
    )
    template = ExcelTemplate([mapping], include_sample_data=True)
    worksheet = _workbook_from_bytes(template.to_bytes())[mapping.sheet_name]

    assert worksheet.cell(row=2, column=1).value == "'=CMD()"


def test_enum_data_validation_skipped_for_long_formula() -> None:
    long_values = [f"value_{index:03d}_long_text" for index in range(30)]
    mapping = ExcelMapping(
        model_class=SimpleUser,
        sheet_name="long_enum",
        columns=[
            ColumnMapping(
                name="status",
                excel_header="Status",
                python_type=str,
                sqla_type=String(),
                nullable=False,
                primary_key=False,
                has_default=False,
                enum_values=long_values,
            )
        ],
        key_columns=[],
    )
    template = ExcelTemplate([mapping])
    worksheet = _workbook_from_bytes(template.to_bytes())[mapping.sheet_name]

    assert len(list(worksheet.data_validations.dataValidation)) == 0
    comment = worksheet.cell(row=1, column=1).comment
    assert comment is not None
    assert "Dropdown omitted" in comment.text
