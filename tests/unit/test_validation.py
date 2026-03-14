from __future__ import annotations

import enum
from datetime import date
from io import BytesIO
from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook
from sqlalchemy import Enum as SAEnum
from sqlalchemy.orm import Mapped, mapped_column
from tests.conftest import Base, Employee, EmployeeStatus, SimpleUser

from sqlalchemy_excel.mapping import ExcelMapping
from sqlalchemy_excel.validation.engine import ExcelValidator
from sqlalchemy_excel.validation.pydantic_backend import PydanticBackend

if TYPE_CHECKING:
    from pathlib import Path


def _create_test_xlsx(
    path: str | Path, headers: list[str], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    _ = worksheet.append(headers)
    for row in rows:
        _ = worksheet.append(row)
    workbook.save(path)


def test_validate_valid_row() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {"id": 1, "name": "Alice", "email": "alice@example.com", "age": 30}

    errors = backend.validate_row(row, row_number=2)

    assert errors == []


def test_validate_null_required_field() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {"id": 1, "name": None, "email": "alice@example.com", "age": 30}

    errors = backend.validate_row(row, row_number=2)

    assert len(errors) == 1
    assert errors[0].column == "name"
    assert errors[0].error_code == "null_error"


def test_validate_wrong_type() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {
        "id": 1,
        "name": "Alice",
        "email": "alice@example.com",
        "age": "not_a_number",
    }

    errors = backend.validate_row(row, row_number=2)

    assert len(errors) == 1
    assert errors[0].column == "age"
    assert errors[0].error_code == "type_error"


def test_validate_nullable_field_none() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {"id": 1, "name": "Alice", "email": "alice@example.com", "age": None}

    errors = backend.validate_row(row, row_number=2)

    assert errors == []


def test_validate_string_max_length() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {
        "id": 1,
        "name": "x" * 200,
        "email": "alice@example.com",
        "age": 30,
    }

    errors = backend.validate_row(row, row_number=2)

    assert len(errors) == 1
    assert errors[0].column == "name"
    assert errors[0].error_code == "length_error"


def test_validate_enum_valid() -> None:
    mapping = ExcelMapping.from_model(Employee)
    backend = PydanticBackend(mapping)
    row = {
        "id": 1,
        "email": "alice@example.com",
        "first_name": "Alice",
        "last_name": "Smith",
        "status": EmployeeStatus.ACTIVE.value,
        "salary": 85000.0,
        "hire_date": date(2024, 1, 1),
        "department_id": None,
        "notes": None,
    }

    errors = backend.validate_row(row, row_number=2)

    assert errors == []


def test_validate_enum_invalid() -> None:
    mapping = ExcelMapping.from_model(Employee)
    backend = PydanticBackend(mapping)
    row = {
        "id": 1,
        "email": "alice@example.com",
        "first_name": "Alice",
        "last_name": "Smith",
        "status": "unknown",
        "salary": 85000.0,
        "hire_date": date(2024, 1, 1),
        "department_id": None,
        "notes": None,
    }

    errors = backend.validate_row(row, row_number=2)

    assert len(errors) == 1
    assert errors[0].column == "status"
    assert errors[0].error_code == "enum_error"


def test_coerce_string_to_int() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    backend = PydanticBackend(mapping)
    row = {"id": 1, "name": "Alice", "email": "alice@example.com", "age": "25"}

    errors = backend.validate_row(row, row_number=2)

    assert errors == []


def test_coerce_string_to_float() -> None:
    mapping = ExcelMapping.from_model(Employee)
    backend = PydanticBackend(mapping)
    row = {
        "id": 1,
        "email": "alice@example.com",
        "first_name": "Alice",
        "last_name": "Smith",
        "status": EmployeeStatus.ACTIVE.value,
        "salary": "85000.0",
        "hire_date": date(2024, 1, 1),
        "department_id": None,
        "notes": None,
    }

    errors = backend.validate_row(row, row_number=2)

    assert errors == []


def test_validate_valid_file(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "valid_users.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[
            [1, "Alice", "alice@example.com", 30],
            [2, "Bob", "bob@example.com", None],
        ],
    )

    report = validator.validate(str(file_path))

    assert report.has_errors is False
    assert report.total_rows == 2
    assert report.valid_rows == 2
    assert report.invalid_rows == 0


def test_validate_invalid_file(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "invalid_users.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[
            [1, None, "alice@example.com", 30],
            [2, "Bob", "bob@example.com", "not_a_number"],
        ],
    )

    report = validator.validate(str(file_path))

    assert report.has_errors is True
    assert len(report.errors) >= 1
    assert report.invalid_rows == 2


def test_validate_max_errors(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "many_invalid_users.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[
            [1, None, "alice@example.com", 30],
            [2, None, "bob@example.com", 35],
            [3, None, "carol@example.com", 40],
        ],
    )

    report = validator.validate(str(file_path), max_errors=2)

    assert report.has_errors is True
    assert len(report.errors) <= 2
    assert len(report.errors) == 2


def test_validate_stop_on_first_error(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "stop_on_first.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[
            [1, None, "alice@example.com", 30],
            [2, None, "bob@example.com", 35],
            [3, "Carol", "carol@example.com", 40],
        ],
    )

    report = validator.validate(str(file_path), stop_on_first_error=True)

    assert report.has_errors is True
    assert report.invalid_rows == 1
    assert len(report.errors) == 1
    assert report.errors[0].row == 2


def test_validate_from_bytesio(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "bytes_users.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[[1, "Alice", "alice@example.com", 25]],
    )

    with file_path.open("rb") as file_obj:
        source = BytesIO(file_obj.read())

    report = validator.validate(source)

    assert report.has_errors is False
    assert report.total_rows == 1


def test_validate_empty_mappings_raises() -> None:
    with pytest.raises(ValueError, match="At least one mapping is required"):
        _ = ExcelValidator([])


def test_imports_from_conftest_for_models() -> None:
    assert Base is not None
    assert SimpleUser is not None
    assert Employee is not None
    assert EmployeeStatus.ACTIVE.value == "active"


def test_validate_header_normalization_matches_reader(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "normalized_special_headers.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["ID", "Name!!!", "Email@", "Age($)"],
        rows=[[1, "Alice", "alice@example.com", 30]],
    )

    report = validator.validate(str(file_path))

    assert report.has_errors is False
    assert report.valid_rows == 1


def test_validate_multiple_mappings_validates_all_sheets(tmp_path: Path) -> None:
    user_mapping = ExcelMapping.from_model(SimpleUser)
    employee_mapping = ExcelMapping.from_model(Employee)

    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = user_mapping.sheet_name
    first.append([column.excel_header for column in user_mapping.columns])
    first.append([1, None, "alice@example.com", 30])

    second = workbook.create_sheet(employee_mapping.sheet_name)
    second.append([column.excel_header for column in employee_mapping.columns])
    second.append(
        [
            1,
            "emp@example.com",
            "Emp",
            "User",
            "not_a_valid_status",
            100.0,
            date(2024, 1, 1),
            None,
            None,
        ]
    )

    file_path = tmp_path / "multi_sheet_validation.xlsx"
    workbook.save(file_path)

    report = ExcelValidator([user_mapping, employee_mapping]).validate(str(file_path))

    assert report.has_errors is True
    assert report.invalid_rows == 2
    assert report.total_rows == 2


class _FiveState(enum.Enum):
    A = "a"
    B = "b"
    C = "c"
    D = "d"
    E = "e"


class _ManyEnumModel(Base):
    __tablename__ = "many_enum_model"

    id: Mapped[int] = mapped_column(primary_key=True)
    status: Mapped[_FiveState] = mapped_column(SAEnum(_FiveState))


def test_validate_enum_enforced_for_more_than_three_values() -> None:
    mapping = ExcelMapping.from_model(_ManyEnumModel)
    backend = PydanticBackend(mapping)

    errors = backend.validate_row({"id": 1, "status": "invalid"}, row_number=2)

    assert len(errors) == 1
    assert errors[0].column == "status"
    assert errors[0].error_code == "enum_error"


def test_validate_unknown_sheet_name_raises_value_error(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])
    file_path = tmp_path / "single_sheet.xlsx"
    _create_test_xlsx(
        str(file_path),
        headers=["id", "name", "email", "age"],
        rows=[[1, "Alice", "alice@example.com", 30]],
    )

    with pytest.raises(ValueError, match="No mapping found for sheet"):
        validator.validate(str(file_path), sheet_name="missing_sheet")


def test_validator_uses_read_only_openpyxl_reader() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    validator = ExcelValidator([mapping])

    assert validator._reader.read_only is True
