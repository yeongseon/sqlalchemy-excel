from __future__ import annotations

# pyright: reportMissingImports=false
from io import BytesIO
from typing import TYPE_CHECKING, cast

import pytest
from openpyxl import Workbook

from sqlalchemy_excel.exceptions import FileFormatError, ReaderError, SheetNotFoundError
from sqlalchemy_excel.reader import OpenpyxlReader
from sqlalchemy_excel.reader.base import normalize_header

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet


def _create_test_xlsx(
    path: Path,
    headers: list[str],
    rows: list[list[object]],
    sheet_name: str = "Sheet1",
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def test_normalize_header_basic() -> None:
    assert normalize_header("First Name") == "first_name"


def test_normalize_header_strips() -> None:
    assert normalize_header("  Name  ") == "name"


def test_normalize_header_special_chars() -> None:
    assert normalize_header("Price ($)") == "price_"


def test_normalize_header_already_normalized() -> None:
    assert normalize_header("email") == "email"


def test_read_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "basic.xlsx"
    _create_test_xlsx(
        path=file_path,
        headers=["Name", "Email"],
        rows=[["Alice", "alice@example.com"], ["Bob", "bob@example.com"]],
    )

    result = OpenpyxlReader().read(file_path)

    assert result.headers == ["name", "email"]
    assert result.total_rows == 2
    assert list(result.rows) == [
        {"name": "Alice", "email": "alice@example.com"},
        {"name": "Bob", "email": "bob@example.com"},
    ]


def test_read_from_path(tmp_path: Path) -> None:
    file_path = tmp_path / "from_path.xlsx"
    _create_test_xlsx(file_path, ["Name"], [["Alice"]])

    result = OpenpyxlReader().read(str(file_path))

    assert result.headers == ["name"]
    assert result.total_rows == 1


def test_read_from_pathlib(tmp_path: Path) -> None:
    file_path = tmp_path / "from_pathlib.xlsx"
    _create_test_xlsx(file_path, ["Email"], [["alice@example.com"]])

    result = OpenpyxlReader().read(file_path)

    assert result.headers == ["email"]
    assert result.total_rows == 1


def test_read_from_bytesio() -> None:
    stream = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["Name", "Email"])
    worksheet.append(["Alice", "alice@example.com"])
    workbook.save(stream)
    workbook.close()
    _ = stream.seek(0)

    result = OpenpyxlReader().read(stream)

    assert result.headers == ["name", "email"]
    assert result.total_rows == 1
    assert list(result.rows) == [{"name": "Alice", "email": "alice@example.com"}]


def test_read_specific_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "multi_sheet.xlsx"
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "First"
    first.append(["Name"])
    first.append(["Alice"])

    second = cast("Worksheet", workbook.create_sheet("Target"))
    second.append(["Name"])
    second.append(["Bob"])

    workbook.save(file_path)
    workbook.close()

    result = OpenpyxlReader().read(file_path, sheet_name="Target")

    assert result.headers == ["name"]
    assert list(result.rows) == [{"name": "Bob"}]


def test_read_missing_sheet_raises(tmp_path: Path) -> None:
    file_path = tmp_path / "missing_sheet.xlsx"
    _create_test_xlsx(file_path, ["Name"], [["Alice"]], sheet_name="Available")

    with pytest.raises(SheetNotFoundError):
        OpenpyxlReader().read(file_path, sheet_name="Missing")


def test_read_invalid_file_raises(tmp_path: Path) -> None:
    file_path = tmp_path / "not_excel.txt"
    _ = file_path.write_text("not an xlsx", encoding="utf-8")

    with pytest.raises(FileFormatError):
        OpenpyxlReader().read(file_path)


def test_read_header_auto_detect(tmp_path: Path) -> None:
    file_path = tmp_path / "auto_header.xlsx"
    _create_test_xlsx(file_path, ["Name", "Email"], [["Alice", "alice@example.com"]])

    result = OpenpyxlReader().read(file_path)

    assert result.headers == ["name", "email"]
    assert list(result.rows) == [{"name": "Alice", "email": "alice@example.com"}]


def test_read_explicit_header_row(tmp_path: Path) -> None:
    file_path = tmp_path / "explicit_header.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.append(["metadata"])
    worksheet.append([""])
    worksheet.append(["Name", "Email"])
    worksheet.append(["Alice", "alice@example.com"])
    workbook.save(file_path)
    workbook.close()

    result = OpenpyxlReader().read(file_path, header_row=3)

    assert result.headers == ["name", "email"]
    assert list(result.rows) == [{"name": "Alice", "email": "alice@example.com"}]


def test_read_skips_empty_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "skip_empty.xlsx"
    _create_test_xlsx(
        file_path,
        ["Name", "Email"],
        [
            ["Alice", "alice@example.com"],
            [None, None],
            ["", "  "],
            ["Bob", "bob@example.com"],
        ],
    )

    result = OpenpyxlReader().read(file_path)

    assert list(result.rows) == [
        {"name": "Alice", "email": "alice@example.com"},
        {"name": "Bob", "email": "bob@example.com"},
    ]
    assert result.total_rows == 2


def test_read_normalizes_headers(tmp_path: Path) -> None:
    file_path = tmp_path / "normalized_headers.xlsx"
    _create_test_xlsx(file_path, ["First Name", "Last Name"], [["Alice", "Smith"]])

    result = OpenpyxlReader().read(file_path)

    assert result.headers == ["first_name", "last_name"]
    assert list(result.rows) == [{"first_name": "Alice", "last_name": "Smith"}]


def test_read_file_size_limit(tmp_path: Path) -> None:
    file_path = tmp_path / "size_limit.xlsx"
    _create_test_xlsx(file_path, ["Name"], [["Alice"]])

    reader = OpenpyxlReader(max_file_size=10)
    with pytest.raises(ReaderError, match="exceeds maximum"):
        reader.read(file_path)


def test_read_empty_worksheet(tmp_path: Path) -> None:
    file_path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    workbook.save(file_path)
    workbook.close()

    with pytest.raises(ReaderError, match="worksheet is empty"):
        OpenpyxlReader().read(file_path)


def test_read_duplicate_headers(tmp_path: Path) -> None:
    file_path = tmp_path / "duplicate_headers.xlsx"
    _create_test_xlsx(file_path, ["First Name", "First_Name"], [["Alice", "A."]])

    with pytest.raises(ReaderError, match="Duplicate normalized header"):
        OpenpyxlReader().read(file_path)


def test_read_returns_dicts(tmp_path: Path) -> None:
    file_path = tmp_path / "dict_rows.xlsx"
    _create_test_xlsx(file_path, ["Name", "Age"], [["Alice", 30], ["Bob", 25]])

    result = OpenpyxlReader().read(file_path)
    rows = list(result.rows)

    assert rows
    assert all(isinstance(row, dict) for row in rows)
    assert rows[0]["name"] == "Alice"
    assert rows[0]["age"] == 30


def test_read_only_mode(tmp_path: Path) -> None:
    file_path = tmp_path / "read_only.xlsx"
    _create_test_xlsx(file_path, ["Name"], [["Alice"], ["Bob"]])

    result = OpenpyxlReader(read_only=True).read(file_path)
    rows = list(result.rows)

    assert result.headers == ["name"]
    assert result.total_rows is None
    assert rows == [{"name": "Alice"}, {"name": "Bob"}]
