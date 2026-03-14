from __future__ import annotations

import importlib
from io import BytesIO
from typing import TYPE_CHECKING

import openpyxl
import pytest

_conftest = importlib.import_module("conftest")
Base = _conftest.Base
Employee = _conftest.Employee
SimpleUser = _conftest.SimpleUser

_sqlalchemy_excel = importlib.import_module("sqlalchemy_excel")
ExcelExporter = _sqlalchemy_excel.ExcelExporter
ExcelMapping = _sqlalchemy_excel.ExcelMapping

_sqlalchemy_excel_exceptions = importlib.import_module("sqlalchemy_excel.exceptions")
ExportError = _sqlalchemy_excel_exceptions.ExportError

if TYPE_CHECKING:
    from pathlib import Path


def test_export_to_file(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])
    output = tmp_path / "users.xlsx"

    result = exporter.export(
        rows=[{"id": 1, "name": "Alice", "email": "alice@example.com", "age": 30}],
        path=output,
    )

    assert result is None
    assert output.exists()
    workbook = openpyxl.load_workbook(output)
    assert workbook.active is not None


def test_export_to_bytes() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])

    content = exporter.export(
        rows=[{"id": 1, "name": "Alice", "email": "alice@example.com", "age": 30}],
        path=None,
    )

    assert isinstance(content, bytes)
    assert len(content) > 0
    workbook = openpyxl.load_workbook(BytesIO(content))
    assert workbook.active is not None


def test_export_header_formatting() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])

    content = exporter.export(rows=[], path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    header = ws["A1"]
    assert header.fill.fill_type == "solid"
    assert header.fill.start_color.rgb is not None
    assert header.fill.start_color.rgb.endswith("4472C4")
    assert header.font.bold is True
    assert header.font.color is not None
    assert header.font.color.rgb is not None
    assert header.font.color.rgb.endswith("FFFFFF")


def test_export_data_values() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])
    rows = [
        {"id": 1, "name": "Alice", "email": "alice@example.com", "age": 30},
        {"id": 2, "name": "Bob", "email": "bob@example.com", "age": 40},
    ]

    content = exporter.export(rows=rows, path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "Alice"
    assert ws.cell(row=2, column=3).value == "alice@example.com"
    assert ws.cell(row=2, column=4).value == 30
    assert ws.cell(row=3, column=1).value == 2
    assert ws.cell(row=3, column=2).value == "Bob"
    assert ws.cell(row=3, column=3).value == "bob@example.com"
    assert ws.cell(row=3, column=4).value == 40


def test_export_auto_filter() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])

    content = exporter.export(
        rows=[
            {"id": 1, "name": "Alice", "email": "alice@example.com", "age": 30},
            {"id": 2, "name": "Bob", "email": "bob@example.com", "age": 40},
        ],
        path=None,
    )
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.auto_filter.ref == "A1:D3"


def test_export_freeze_panes() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])

    content = exporter.export(rows=[], path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.freeze_panes == "A2"


def test_export_from_orm_objects(session, sample_departments, sample_employees) -> None:
    assert issubclass(Employee, Base)
    assert sample_departments

    mapping = ExcelMapping.from_model(
        Employee,
        include=["id", "email", "first_name", "last_name", "salary"],
    )
    exporter = ExcelExporter([mapping])

    content = exporter.export(rows=sample_employees, path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "alice@example.com"
    assert ws.cell(row=2, column=3).value == "Alice"


def test_export_from_dicts() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])
    rows = [
        {"id": 1, "name": "First", "email": "first@example.com", "age": 20},
        {"id": 2, "name": "Second", "email": "second@example.com", "age": None},
    ]

    content = exporter.export(rows=rows, path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.cell(row=2, column=2).value == "First"
    assert ws.cell(row=3, column=2).value == "Second"


def test_export_empty_mappings_raises() -> None:
    with pytest.raises(ExportError, match="At least one ExcelMapping is required"):
        _ = ExcelExporter([])


def test_export_empty_rows() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    exporter = ExcelExporter([mapping])

    content = exporter.export(rows=[], path=None)
    assert isinstance(content, bytes)

    ws = openpyxl.load_workbook(BytesIO(content)).active
    assert ws is not None
    assert ws.max_row == 1
    assert ws.cell(row=1, column=1).value == "Id"
    assert ws.cell(row=1, column=2).value == "Name"
    assert ws.cell(row=1, column=3).value == "Email"
    assert ws.cell(row=1, column=4).value == "Age"
