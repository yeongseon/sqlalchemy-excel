from __future__ import annotations

# pyright: reportMissingImports=false, reportUnknownVariableType=false, reportUnknownMemberType=false, reportUnknownArgumentType=false, reportUnknownParameterType=false, reportMissingParameterType=false
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from tests.conftest import Base, Department, SimpleUser

from sqlalchemy_excel import (
    ExcelExporter,
    ExcelImporter,
    ExcelMapping,
    ExcelTemplate,
    ExcelValidator,
    ValidationReport,
)

if TYPE_CHECKING:
    from pathlib import Path

CellValue = int | float | str | bool | None


def _write_rows(path: Path, sheet_name: str, rows: list[tuple[CellValue, ...]]) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    for row_index, row_values in enumerate(rows, start=2):
        for column_index, value in enumerate(row_values, start=1):
            _ = worksheet.cell(row=row_index, column=column_index, value=value)
    workbook.save(path)


def test_full_pipeline_simple_user(engine, session: Session, tmp_path: Path) -> None:
    assert engine is not None

    mapping = ExcelMapping.from_model(SimpleUser)
    template_path = tmp_path / "simple_user_template.xlsx"
    export_path = tmp_path / "simple_user_export.xlsx"

    ExcelTemplate([mapping]).save(template_path)
    _write_rows(
        template_path,
        mapping.sheet_name,
        [
            (1, "Alice", "alice@test.com", 30),
            (2, "Bob", "bob@test.com", 25),
            (3, "Charlie", "charlie@test.com", 40),
        ],
    )

    report = ExcelValidator([mapping]).validate(template_path)
    assert isinstance(report, ValidationReport)
    assert report.has_errors is False

    result = ExcelImporter([mapping], session).insert(template_path, validate=False)
    assert result.inserted == 3
    session.commit()

    assert session.query(SimpleUser).count() == 3

    rows = session.query(SimpleUser).order_by(SimpleUser.id).all()
    ExcelExporter([mapping]).export(rows, export_path)

    assert export_path.exists()
    exported = load_workbook(export_path)
    exported_sheet = exported[mapping.sheet_name]
    assert exported_sheet.max_row == 4


def test_pipeline_validation_catches_errors(tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    template_path = tmp_path / "simple_user_invalid.xlsx"

    ExcelTemplate([mapping]).save(template_path)
    _write_rows(
        template_path,
        mapping.sheet_name,
        [
            (1, None, "missing-name@test.com", 30),
        ],
    )

    report = ExcelValidator([mapping]).validate(template_path)
    assert report.has_errors is True
    assert report.invalid_rows == 1
    assert report.errors

    first_error = report.errors[0]
    assert first_error.column == "name"
    assert first_error.error_code == "null_error"


def test_pipeline_template_to_bytesio() -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    stream = ExcelTemplate([mapping]).to_bytesio()

    assert isinstance(stream, BytesIO)

    workbook = load_workbook(stream)
    worksheet = workbook[mapping.sheet_name]
    actual_headers = [
        worksheet.cell(row=1, column=index).value for index in range(1, 5)
    ]
    expected_headers = [column.excel_header for column in mapping.columns]
    assert actual_headers == expected_headers

    local_engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(local_engine)
    with Session(local_engine) as local_session:
        assert local_session.query(SimpleUser).count() == 0


def test_pipeline_multiple_models(tmp_path: Path) -> None:
    user_mapping = ExcelMapping.from_model(SimpleUser)
    department_mapping = ExcelMapping.from_model(Department)
    template_path = tmp_path / "multiple_models_template.xlsx"

    ExcelTemplate([user_mapping, department_mapping]).save(template_path)
    workbook = load_workbook(template_path)

    assert workbook.sheetnames == [
        user_mapping.sheet_name,
        department_mapping.sheet_name,
    ]
    assert len(workbook.worksheets) == 2


def test_pipeline_upsert(session: Session, tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    importer = ExcelImporter([mapping], session)

    initial_path = tmp_path / "upsert_initial.xlsx"
    ExcelTemplate([mapping]).save(initial_path)
    _write_rows(
        initial_path,
        mapping.sheet_name,
        [
            (1, "Alice", "alice@test.com", 30),
            (2, "Bob", "bob@test.com", 25),
        ],
    )
    initial_result = importer.insert(initial_path, validate=False)
    assert initial_result.inserted == 2
    session.commit()

    upsert_path = tmp_path / "upsert_update.xlsx"
    ExcelTemplate([mapping]).save(upsert_path)
    _write_rows(
        upsert_path,
        mapping.sheet_name,
        [
            (1, "Alice Updated", "alice@test.com", 30),
            (3, "Charlie", "charlie@test.com", 40),
        ],
    )

    upsert_result = importer.upsert(upsert_path, validate=False)
    assert upsert_result.inserted == 1
    assert upsert_result.updated == 1
    session.commit()

    assert session.query(SimpleUser).count() == 3
    updated_user = session.query(SimpleUser).filter_by(id=1).one()
    assert updated_user.name == "Alice Updated"


def test_pipeline_dry_run(session: Session, tmp_path: Path) -> None:
    mapping = ExcelMapping.from_model(SimpleUser)
    path = tmp_path / "dry_run.xlsx"

    ExcelTemplate([mapping]).save(path)
    _write_rows(
        path,
        mapping.sheet_name,
        [
            (1, "Alice", "alice@test.com", 30),
            (2, "Bob", "bob@test.com", 25),
        ],
    )

    result = ExcelImporter([mapping], session).dry_run(path, validate=False)
    assert result.inserted == 2
    assert session.query(SimpleUser).count() == 0
